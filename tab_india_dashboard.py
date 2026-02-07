"""
Tab: India Intelligence Dashboard - VIDEO OPTIMIZED
Clean, professional visualizations perfect for screen recording and screenshots
Focused exclusively on India data - no USA comparison
Enhanced with word clouds, cross-platform analysis, and comprehensive analytics
"""
import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
from datetime import date, timedelta
from collections import Counter
from difflib import SequenceMatcher
from wordcloud import WordCloud
import matplotlib.pyplot as plt
import os


def render_india_intelligence_dashboard(supabase):
    """
    India-only intelligence dashboard optimized for video content
    Clean, professional charts ready for recording
    """
    
    # Header with India flag colors
    st.markdown(
        """<div style="background:linear-gradient(to right, #ff9933, #ffffff, #138808); 
        padding:30px; border-radius:15px; text-align:center; margin-bottom:30px;">
        <h1 style="color:#000080; margin:0;">🇮🇳 India Intelligence Dashboard</h1>
        <p style="color:#000080; font-size:18px; margin:10px 0 0 0;">
        Real-time trend analysis for content creators
        </p>
        </div>""",
        unsafe_allow_html=True
    )
    
    if not supabase:
        st.error("❌ Supabase not configured")
        return
    
    # Date selector for flexibility
    col_date, col_refresh = st.columns([3, 1])
    
    with col_date:
        selected_date = st.date_input(
            "📅 Analysis Date",
            value=date.today(),
            max_value=date.today(),
            key='india_date'
        )
    
    with col_refresh:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🔄 Refresh Data", use_container_width=True):
            st.rerun()
    
    # Fetch India intelligence
    intelligence = fetch_india_intelligence(supabase, selected_date.isoformat())
    
    if not intelligence:
        st.warning(f"⚠️ No intelligence data found for {selected_date.isoformat()}")
        st.info("💡 Generate intelligence in the **Intelligence Analysis** tab first")
        return
    
    # Fetch trend data for analytics
    google_data, twitter_data = fetch_india_trends(supabase, selected_date.isoformat())
    
    st.divider()
    
    # === SECTION 1: KEY METRICS (Perfect for video intro) ===
    render_key_metrics_section(intelligence)
    
    st.divider()
    
    # === SECTION 2: TRENDING THEMES (Main content) ===
    render_trending_themes_section(intelligence)
    
    st.divider()
    
    # === SECTION 3: ANOMALY DETECTION (Highlight reel) ===
    render_anomalies_section(intelligence)
    
    st.divider()
    
    # === SECTION 4: WORD CLOUDS (Visual appeal) ===
    if google_data or twitter_data:
        render_wordclouds_section(google_data, twitter_data)
        st.divider()
    
    # === SECTION 5: CROSS-PLATFORM ANALYSIS ===
    if google_data and twitter_data:
        render_india_cross_platform_analysis(google_data, twitter_data)
        st.divider()
    
    # === SECTION 6: VIRAL TRENDS ===
    if google_data or twitter_data:
        render_india_viral_trends(google_data, twitter_data)
        st.divider()
    
    # === SECTION 7: CATEGORY BREAKDOWN (3 charts) ===
    if google_data or twitter_data:
        render_category_breakdown_three_charts(google_data, twitter_data)
        st.divider()
    
    # === SECTION 8: TOP TRENDS WITH CONTEXT ===
    if google_data or twitter_data:
        render_top_trends_with_context_india(google_data, twitter_data)
        st.divider()
    
    # === SECTION 9: SENTIMENT ANALYSIS ===
    render_sentiment_analysis(intelligence)
    
    st.divider()
    
    # === SECTION 10: SENTIMENT DISTRIBUTION WHEEL ===
    if google_data or twitter_data:
        render_sentiment_distribution_wheel_india(google_data, twitter_data)
        st.divider()
    
    # === SECTION 11: CANVA EXPORT (Excel download) ===
    render_excel_export_section(supabase, selected_date.isoformat())
    
    st.divider()
    
    # === SECTION 12: RAW DATA TABLES ===
    if google_data or twitter_data:
        render_raw_data_tables_india(google_data, twitter_data)


def fetch_india_intelligence(supabase, analysis_date):
    """Fetch India intelligence for specific date"""
    try:
        result = supabase.table('daily_insights')\
            .select('*')\
            .eq('analysis_date', analysis_date)\
            .eq('region', 'India')\
            .execute()
        
        if result.data and len(result.data) > 0:
            return result.data[0]
        return None
    except Exception as e:
        st.error(f"Error fetching intelligence: {str(e)}")
        return None


def fetch_india_trends(supabase, analysis_date):
    """Fetch India trend data for analytics"""
    try:
        google_result = supabase.table('google_trends')\
            .select('*')\
            .eq('collection_date', analysis_date)\
            .eq('region', 'India')\
            .execute()
        
        twitter_result = supabase.table('twitter_trends')\
            .select('*')\
            .eq('collection_date', analysis_date)\
            .eq('region', 'India')\
            .execute()
        
        google_data = google_result.data if google_result.data else []
        twitter_data = twitter_result.data if twitter_result.data else []
        
        return google_data, twitter_data
    except Exception as e:
        st.error(f"Error fetching trends: {str(e)}")
        return [], []


def render_key_metrics_section(intelligence):
    """
    Top-line metrics in large, video-friendly cards
    Perfect for opening shots
    """
    st.markdown("### 📊 Today's Intelligence Summary")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        theme_1_title = intelligence.get('theme_1_title', 'N/A')
        theme_1_category = intelligence.get('theme_1_category', 'Unknown')
        
        st.markdown(
            f"""<div style="background:linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
            padding:30px; border-radius:15px; color:white; text-align:center; min-height:200px;">
            <div style="font-size:48px; margin-bottom:15px;">🔥</div>
            <div style="font-size:14px; opacity:0.9; margin-bottom:10px;">TOP THEME</div>
            <div style="font-size:22px; font-weight:700; margin-bottom:15px;">{theme_1_title}</div>
            <div style="font-size:16px; opacity:0.8;">{theme_1_category}</div>
            </div>""",
            unsafe_allow_html=True
        )
    
    with col2:
        anomaly_1_keyword = intelligence.get('anomaly_1_keyword', 'N/A')
        anomaly_1_velocity = intelligence.get('anomaly_1_velocity', 'N/A')
        
        st.markdown(
            f"""<div style="background:linear-gradient(135deg, #f093fb 0%, #f5576c 100%); 
            padding:30px; border-radius:15px; color:white; text-align:center; min-height:200px;">
            <div style="font-size:48px; margin-bottom:15px;">⚡</div>
            <div style="font-size:14px; opacity:0.9; margin-bottom:10px;">BREAKOUT TREND</div>
            <div style="font-size:22px; font-weight:700; margin-bottom:15px;">{anomaly_1_keyword}</div>
            <div style="font-size:16px; opacity:0.8;">{anomaly_1_velocity.upper()}</div>
            </div>""",
            unsafe_allow_html=True
        )
    
    with col3:
        sentiment = intelligence.get('overall_sentiment', 0)
        sentiment_pct = int(sentiment * 100) if sentiment else 50
        
        if sentiment_pct >= 60:
            mood_emoji = "😊"
            mood_text = "POSITIVE"
            mood_color = "#10b981"
        elif sentiment_pct <= 40:
            mood_emoji = "😟"
            mood_text = "CONCERNED"
            mood_color = "#ef4444"
        else:
            mood_emoji = "😐"
            mood_text = "NEUTRAL"
            mood_color = "#f59e0b"
        
        st.markdown(
            f"""<div style="background:linear-gradient(135deg, {mood_color} 0%, {mood_color}dd 100%); 
            padding:30px; border-radius:15px; color:white; text-align:center; min-height:200px;">
            <div style="font-size:48px; margin-bottom:15px;">{mood_emoji}</div>
            <div style="font-size:14px; opacity:0.9; margin-bottom:10px;">NATIONAL MOOD</div>
            <div style="font-size:22px; font-weight:700; margin-bottom:15px;">{mood_text}</div>
            <div style="font-size:16px; opacity:0.8;">{sentiment_pct}% Index</div>
            </div>""",
            unsafe_allow_html=True
        )


def render_trending_themes_section(intelligence):
    """
    Visual breakdown of top 2 themes
    Great for detailed analysis segments
    """
    st.markdown("### 🎯 Top Trending Themes")
    
    col_theme1, col_theme2 = st.columns(2)
    
    with col_theme1:
        st.markdown("#### 🥇 Theme #1")
        
        theme_data = {
            'Title': intelligence.get('theme_1_title', 'N/A'),
            'Category': intelligence.get('theme_1_category', 'N/A'),
            'Mood': intelligence.get('theme_1_mood', 'N/A'),
            'Keywords': ', '.join(intelligence.get('theme_1_keywords', [])) if intelligence.get('theme_1_keywords') else 'N/A'
        }
        
        for key, value in theme_data.items():
            st.markdown(
                f"""<div style="background:#1e293b; padding:15px; border-radius:10px; 
                margin-bottom:10px; border-left:4px solid #3b82f6;">
                <strong style="color:#60a5fa;">{key}:</strong> 
                <span style="color:#e2e8f0;">{value}</span>
                </div>""",
                unsafe_allow_html=True
            )
        
        with st.expander("📖 Full Context", expanded=False):
            st.info(intelligence.get('theme_1_context', 'N/A'))
            st.success(f"**Deep Why:** {intelligence.get('theme_1_deep_why', 'N/A')}")
            st.warning(f"**Big Question:** {intelligence.get('theme_1_big_question', 'N/A')}")
    
    with col_theme2:
        st.markdown("#### 🥈 Theme #2")
        
        theme_data = {
            'Title': intelligence.get('theme_2_title', 'N/A'),
            'Category': intelligence.get('theme_2_category', 'N/A'),
            'Mood': intelligence.get('theme_2_mood', 'N/A'),
            'Keywords': ', '.join(intelligence.get('theme_2_keywords', [])) if intelligence.get('theme_2_keywords') else 'N/A'
        }
        
        for key, value in theme_data.items():
            st.markdown(
                f"""<div style="background:#1e293b; padding:15px; border-radius:10px; 
                margin-bottom:10px; border-left:4px solid #8b5cf6;">
                <strong style="color:#a78bfa;">{key}:</strong> 
                <span style="color:#e2e8f0;">{value}</span>
                </div>""",
                unsafe_allow_html=True
            )
        
        with st.expander("📖 Full Context", expanded=False):
            st.info(intelligence.get('theme_2_context', 'N/A'))
            st.success(f"**Deep Why:** {intelligence.get('theme_2_deep_why', 'N/A')}")
            st.warning(f"**Big Question:** {intelligence.get('theme_2_big_question', 'N/A')}")


def render_anomalies_section(intelligence):
    """
    Highlight anomalies with visual impact
    Perfect for "surprise factor" segments
    """
    st.markdown("### 🚨 Anomaly Detection")
    
    col1, col2 = st.columns(2)
    
    with col1:
        keyword = intelligence.get('anomaly_1_keyword', 'N/A')
        velocity = intelligence.get('anomaly_1_velocity', 'N/A')
        explanation = intelligence.get('anomaly_1_explanation', 'N/A')
        
        st.markdown(
            f"""<div style="background:linear-gradient(135deg, #f59e0b 0%, #ef4444 100%); 
            padding:25px; border-radius:12px; color:white;">
            <h3 style="margin:0 0 15px 0;">⚡ Anomaly #1</h3>
            <h2 style="margin:0 0 10px 0;">{keyword}</h2>
            <div style="font-size:18px; opacity:0.9;">Velocity: {velocity.upper()}</div>
            </div>""",
            unsafe_allow_html=True
        )
        
        st.markdown("<br>", unsafe_allow_html=True)
        st.info(f"**Why It Matters:** {explanation}")
    
    with col2:
        keyword = intelligence.get('anomaly_2_keyword', 'N/A')
        velocity = intelligence.get('anomaly_2_velocity', 'N/A')
        explanation = intelligence.get('anomaly_2_explanation', 'N/A')
        
        st.markdown(
            f"""<div style="background:linear-gradient(135deg, #06b6d4 0%, #3b82f6 100%); 
            padding:25px; border-radius:12px; color:white;">
            <h3 style="margin:0 0 15px 0;">⚡ Anomaly #2</h3>
            <h2 style="margin:0 0 10px 0;">{keyword}</h2>
            <div style="font-size:18px; opacity:0.9;">Velocity: {velocity.upper()}</div>
            </div>""",
            unsafe_allow_html=True
        )
        
        st.markdown("<br>", unsafe_allow_html=True)
        st.info(f"**Why It Matters:** {explanation}")


def setup_font():
    """Setup font for Hindi/Devanagari support"""
    try:
        font_path = os.path.abspath('fonts/NotoSansDevanagari-VariableFont_wdth,wght.ttf')
        if os.path.exists(font_path):
            from matplotlib import font_manager
            font_manager.fontManager.addfont(font_path)
            return font_path
    except Exception:
        pass
    
    possible_fonts = [
        '/usr/share/fonts/truetype/noto/NotoSans-Regular.ttf',
        '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
        'C:\\Windows\\Fonts\\NotoSans-Regular.ttf',
        'C:\\Windows\\Fonts\\Arial.ttf',
        '/System/Library/Fonts/Supplemental/Arial Unicode.ttf'
    ]
    
    for font in possible_fonts:
        if os.path.exists(font):
            return font
    
    return None


def render_wordclouds_section(google_data, twitter_data):
    """
    4 word clouds: Google keywords, Twitter keywords, Google categories, Twitter categories
    """
    st.markdown("### ☁️ Word Cloud Visualizations")
    st.caption("💡 Larger text = Higher volume/frequency • Darker color = More prominent")
    
    font_path = setup_font()
    
    # Row 1: Keywords
    st.markdown("#### 📝 Keywords")
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("**Google Search Keywords**")
        fig = create_keyword_wordcloud(google_data, font_path, 'Google')
        if fig:
            st.pyplot(fig)
        else:
            st.info("No Google keyword data available")
    
    with col2:
        st.markdown("**Twitter Trending Topics**")
        fig = create_keyword_wordcloud(twitter_data, font_path, 'Twitter')
        if fig:
            st.pyplot(fig)
        else:
            st.info("No Twitter keyword data available")
    
    st.markdown("---")
    
    # Row 2: Categories
    st.markdown("#### 📊 Categories")
    col3, col4 = st.columns(2)
    
    with col3:
        st.markdown("**Google Search Categories**")
        fig, cat_text = create_category_wordcloud(google_data, font_path, 'Google')
        if fig:
            st.pyplot(fig)
            with st.expander("📊 Category Breakdown", expanded=False):
                st.text(cat_text)
        else:
            st.info("No Google category data available")
    
    with col4:
        st.markdown("**Twitter Trend Categories**")
        fig, cat_text = create_category_wordcloud(twitter_data, font_path, 'Twitter')
        if fig:
            st.pyplot(fig)
            with st.expander("📊 Category Breakdown", expanded=False):
                st.text(cat_text)
        else:
            st.info("No Twitter category data available")


def create_keyword_wordcloud(data, font_path, platform='Google'):
    """Create word cloud from keywords"""
    try:
        if not data or len(data) == 0:
            return None
        
        df = pd.DataFrame(data)
        
        if platform == 'Google':
            word_freq = {row['keyword']: row['search_volume'] for _, row in df.iterrows()}
        else:
            word_freq = {row['keyword']: row['mention_volume'] for _, row in df.iterrows()}
        
        if not word_freq:
            return None
        
        # India orange gradient color function
        def color_func(word, font_size, position, orientation, random_state=None, **kwargs):
            volume = word_freq.get(word, 0)
            max_vol = max(word_freq.values()) if word_freq else 1
            lightness = 70 - (40 * (volume / max_vol))
            return f"hsl(30, 100%, {int(lightness)}%)"
        
        wc = WordCloud(
            width=600,
            height=300,
            background_color='white',
            relative_scaling=0.5,
            min_font_size=10,
            color_func=color_func,
            font_path=font_path,
            collocations=False
        ).generate_from_frequencies(word_freq)
        
        fig, ax = plt.subplots(figsize=(8, 4))
        ax.imshow(wc, interpolation='bilinear')
        ax.axis('off')
        
        return fig
    except Exception as e:
        st.error(f"Error creating keyword wordcloud: {str(e)}")
        return None


def create_category_wordcloud(data, font_path, platform='Google'):
    """Create word cloud from categories"""
    try:
        if not data or len(data) == 0:
            return None, ""
        
        df = pd.DataFrame(data)
        category_counts = Counter(df['category'])
        
        if not category_counts:
            return None, ""
        
        # India orange gradient color function
        def color_func(word, font_size, position, orientation, random_state=None, **kwargs):
            count = category_counts.get(word, 0)
            max_count = max(category_counts.values()) if category_counts else 1
            lightness = 70 - (40 * (count / max_count))
            return f"hsl(30, 100%, {int(lightness)}%)"
        
        wc = WordCloud(
            width=600,
            height=300,
            background_color='white',
            relative_scaling=0.5,
            min_font_size=10,
            color_func=color_func,
            font_path=font_path,
            collocations=False
        ).generate_from_frequencies(category_counts)
        
        fig, ax = plt.subplots(figsize=(8, 4))
        ax.imshow(wc, interpolation='bilinear')
        ax.axis('off')
        
        categories_text = "\n".join([f"{cat}: {count} trends" for cat, count in category_counts.most_common(10)])
        
        return fig, categories_text
    except Exception as e:
        st.error(f"Error creating category wordcloud: {str(e)}")
        return None, ""


def fuzzy_match(str1, str2, threshold=0.8):
    """Check if two strings are similar using fuzzy matching"""
    return SequenceMatcher(None, str1.lower(), str2.lower()).ratio() >= threshold


def render_india_cross_platform_analysis(google_data, twitter_data):
    """Cross-platform overlap analysis for India"""
    st.markdown("### 🔄 Cross-Platform Analysis")
    st.caption("Discover trends appearing on both Google Search and Twitter")
    
    google_df = pd.DataFrame(google_data)
    twitter_df = pd.DataFrame(twitter_data)
    
    google_kw = set(google_df['keyword'].str.lower())
    twitter_kw = set(twitter_df['keyword'].str.lower())
    
    overlaps = []
    for g_kw in google_kw:
        for t_kw in twitter_kw:
            if fuzzy_match(g_kw, t_kw, threshold=0.8):
                overlaps.append((g_kw, t_kw))
                break
    
    st.markdown(f"""
    <div style="background: linear-gradient(135deg, #ea580c 0%, #f59e0b 100%); 
                padding: 20px; border-radius: 12px; color: white; text-align: center; margin-bottom: 20px;">
        <div style="font-size: 3rem; font-weight: 700; margin-bottom: 5px;">{len(overlaps)}</div>
        <div style="font-size: 1.1rem; opacity: 0.95;">Cross-Platform Trends</div>
        <div style="font-size: 0.9rem; opacity: 0.8; margin-top: 5px;">
            Appearing on both Google Search & Twitter
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    if overlaps:
        with st.expander("🔍 View Cross-Platform Trends", expanded=True):
            for g_kw, t_kw in overlaps[:15]:
                if g_kw == t_kw:
                    st.markdown(f"• **{g_kw.title()}**")
                else:
                    st.markdown(f"• **{g_kw.title()}** ≈ {t_kw.title()}")
        
        if len(overlaps) > 10:
            st.success("🎯 **High Synergy:** Strong cross-platform momentum today!")
        elif len(overlaps) > 5:
            st.info("📊 **Moderate Synergy:** Some overlapping themes across platforms")
        else:
            st.warning("🌐 **Low Synergy:** Different interests across platforms")
    else:
        st.info("No significant cross-platform overlaps detected today")


def calculate_viral_score(volume, velocity, sentiment):
    """Calculate viral coefficient score"""
    velocity_multipliers = {
        'breakout': 3.0,
        'spike': 3.0,
        'rising': 2.0,
        'rising_fast': 2.5,
        'high': 2.0,
        'steady': 1.0,
        'moderate': 1.0,
        'slow': 0.5,
        'declining': 0.3
    }
    
    sentiment_boost = {
        'excited': 1.1,
        'celebrating': 1.1,
        'controversial': 1.05,
        'concerned': 1.05,
        'curious': 1.0
    }
    
    velocity_key = str(velocity).lower()
    sentiment_key = str(sentiment).lower()
    
    base_score = (volume / 1000) * velocity_multipliers.get(velocity_key, 1.0)
    boosted_score = base_score * sentiment_boost.get(sentiment_key, 1.0)
    
    return min(100, int(boosted_score / 50))


def render_india_viral_trends(google_data, twitter_data):
    """Display top viral trends for India"""
    st.markdown("### 🔥 Viral Trends")
    st.caption("Trends with highest viral coefficient scores")
    
    viral_trends = []
    
    google_df = pd.DataFrame(google_data)
    twitter_df = pd.DataFrame(twitter_data)
    
    for _, row in google_df.iterrows():
        if 'velocity' in google_df.columns and 'public_sentiment' in google_df.columns:
            score = calculate_viral_score(
                row['search_volume'],
                row.get('velocity', 'steady'),
                row.get('public_sentiment', 'curious')
            )
            viral_trends.append({
                'keyword': row['keyword'],
                'platform': 'Google',
                'volume': row['search_volume'],
                'score': score
            })
    
    for _, row in twitter_df.iterrows():
        if 'velocity' in twitter_df.columns:
            sentiment = row.get('primary_sentiment') or row.get('sentiment', 'curious')
            score = calculate_viral_score(
                row['mention_volume'],
                row.get('velocity', 'steady'),
                sentiment
            )
            viral_trends.append({
                'keyword': row['keyword'],
                'platform': 'Twitter',
                'volume': row['mention_volume'],
                'score': score
            })
    
    if viral_trends:
        viral_df = pd.DataFrame(viral_trends).sort_values('score', ascending=False).head(10)
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("**🔍 Google Top Viral**")
            google_viral = viral_df[viral_df['platform'] == 'Google'].head(5)
            for idx, trend in google_viral.iterrows():
                st.markdown(f"""
                <div style="background: #fff7ed; padding: 12px; border-left: 4px solid #ea580c; 
                           margin-bottom: 10px; border-radius: 6px;">
                    <div style="font-weight: bold; color: #9a3412; font-size: 15px;">
                        {trend['keyword'][:50]}
                    </div>
                    <div style="font-size: 0.9rem; color: #ea580c; margin-top: 5px;">
                        Score: {trend['score']}/100 • {int(trend['volume']/1000)}K searches
                    </div>
                </div>
                """, unsafe_allow_html=True)
        
        with col2:
            st.markdown("**🐦 Twitter Top Viral**")
            twitter_viral = viral_df[viral_df['platform'] == 'Twitter'].head(5)
            for idx, trend in twitter_viral.iterrows():
                st.markdown(f"""
                <div style="background: #fef3c7; padding: 12px; border-left: 4px solid #f59e0b; 
                           margin-bottom: 10px; border-radius: 6px;">
                    <div style="font-weight: bold; color: #92400e; font-size: 15px;">
                        {trend['keyword'][:50]}
                    </div>
                    <div style="font-size: 0.9rem; color: #f59e0b; margin-top: 5px;">
                        Score: {trend['score']}/100 • {int(trend['volume']/1000)}K mentions
                    </div>
                </div>
                """, unsafe_allow_html=True)
    else:
        st.info("No viral trend data available")


def render_category_breakdown_three_charts(google_data, twitter_data):
    """3 separate category distribution charts"""
    st.markdown("### 📊 Category Distribution")
    st.caption("💡 Darker colors indicate higher volume")
    
    google_df = pd.DataFrame(google_data)
    twitter_df = pd.DataFrame(twitter_data)
    
    # Chart 1: Google Categories
    st.markdown("#### 🔍 Google Search Categories")
    if len(google_df) > 0:
        fig = create_category_treemap(google_df, 'Google', 'search_volume')
        if fig:
            st.plotly_chart(fig, use_container_width=True, key='category_google_treemap')
    else:
        st.info("No Google data available")
    
    st.markdown("---")
    
    # Chart 2: Twitter Categories
    st.markdown("#### 🐦 Twitter Trend Categories")
    if len(twitter_df) > 0:
        fig = create_category_treemap(twitter_df, 'Twitter', 'mention_volume')
        if fig:
            st.plotly_chart(fig, use_container_width=True, key='category_twitter_treemap')
    else:
        st.info("No Twitter data available")
    
    st.markdown("---")
    
    # Chart 3: Overall Combined
    st.markdown("#### 📈 Overall Category Distribution")
    if len(google_df) > 0 or len(twitter_df) > 0:
        combined = pd.concat([
            google_df[['category', 'keyword', 'search_volume']].rename(columns={'search_volume': 'volume'}),
            twitter_df[['category', 'keyword', 'mention_volume']].rename(columns={'mention_volume': 'volume'})
        ])
        fig = create_category_treemap(combined, 'Overall', 'volume')
        if fig:
            st.plotly_chart(fig, use_container_width=True, key='category_overall_treemap')
    else:
        st.info("No data available")


def create_category_treemap(df, platform, volume_col):
    """Create treemap for category distribution"""
    try:
        cat_data = []
        for category, group in df.groupby('category'):
            keywords = group['keyword'].tolist()[:10]
            total_volume = group[volume_col].sum()
            
            cat_data.append({
                'category': category,
                'volume': total_volume,
                'keywords': '<br>'.join([f"  • {kw[:40]}" for kw in keywords]),
                'count': len(group)
            })
        
        cat_df = pd.DataFrame(cat_data)
        
        # India orange gradient
        color_scale = ['#fed7aa', '#fdba74', '#fb923c', '#f97316', '#ea580c', '#c2410c', '#9a3412']
        
        fig = px.treemap(
            cat_df,
            path=['category'],
            values='volume',
            color='volume',
            color_continuous_scale=color_scale,
            custom_data=['keywords', 'count'],
            title=f"{platform} Category Distribution"
        )
        
        fig.update_traces(
            textinfo="label+value",
            textfont_size=14,
            hovertemplate='<b>%{label}</b><br>' +
                         'Volume: %{value:,.0f}<br>' +
                         'Trends: %{customdata[1]}<br><br>' +
                         '<b>Top Keywords:</b><br>%{customdata[0]}<extra></extra>',
            marker=dict(
                line=dict(width=2, color='white'),
                pad=dict(t=50, l=5, r=5, b=5)
            )
        )
        
        fig.update_layout(
            height=400,
            title_font_size=18,
            title_x=0.5,
            coloraxis_showscale=False,
            margin=dict(l=10, r=10, t=50, b=10)
        )
        
        return fig
    except Exception as e:
        st.error(f"Error creating treemap: {str(e)}")
        return None


def render_top_trends_with_context_india(google_data, twitter_data):
    """Top trends with context in hover"""
    st.markdown("### 📊 Top Trends with Context")
    st.caption("💡 Hover over bars to see full context and why trending")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("**🔍 Google Top Searches**")
        fig = create_trends_bar_chart(google_data, 'Google')
        if fig:
            st.plotly_chart(fig, use_container_width=True, key='top_trends_google')
        else:
            st.info("No Google data available")
    
    with col2:
        st.markdown("**🐦 Twitter Top Trends**")
        fig = create_trends_bar_chart(twitter_data, 'Twitter')
        if fig:
            st.plotly_chart(fig, use_container_width=True, key='top_trends_twitter')
        else:
            st.info("No Twitter data available")


def create_trends_bar_chart(data, platform='Google'):
    """Create horizontal bar chart with context on hover"""
    try:
        if not data or len(data) == 0:
            return None
        
        df = pd.DataFrame(data)
        
        if platform == 'Google':
            top_data = df.nlargest(10, 'search_volume')
            volume_col = 'search_volume'
        else:
            top_data = df.nlargest(10, 'mention_volume')
            volume_col = 'mention_volume'
        
        top_data = top_data.copy()
        top_data['keyword_short'] = top_data['keyword'].apply(
            lambda x: x[:30] + '...' if len(x) > 30 else x
        )
        
        fig = go.Figure()
        
        fig.add_trace(go.Bar(
            y=top_data['keyword_short'][::-1],
            x=top_data[volume_col][::-1],
            orientation='h',
            marker=dict(color='#ff9933', opacity=0.8),
            text=[f'{int(x/1000)}K' if x < 1000000 else f'{int(x/1000000*10)/10}M'
                  for x in top_data[volume_col][::-1]],
            textposition='outside',
            customdata=top_data[['keyword', 'context', 'why_trending', volume_col]][::-1].values,
            hovertemplate='<b>%{customdata[0]}</b><br>' +
                         'Volume: %{customdata[3]:,.0f}<br><br>' +
                         '<b>Context:</b> %{customdata[1]}<br><br>' +
                         '<b>Why Trending:</b> %{customdata[2]}<extra></extra>',
            showlegend=False
        ))
        
        fig.update_layout(
            height=400,
            margin=dict(l=0, r=50, t=30, b=0),
            paper_bgcolor='white',
            plot_bgcolor='#f9fafb',
            xaxis=dict(showgrid=True, gridcolor='#e5e7eb'),
            yaxis=dict(showgrid=False),
            title=dict(text=f"Top 10 {platform} Trends", font=dict(size=14))
        )
        
        return fig
    except Exception as e:
        st.error(f"Error creating bar chart: {str(e)}")
        return None


def render_sentiment_analysis(intelligence):
    """
    Sentiment gauge meter
    Perfect for mood visualization
    """
    st.markdown("### 🎭 Sentiment Analysis")
    
    sentiment = intelligence.get('overall_sentiment', 0.5)
    sentiment_pct = int(sentiment * 100)
    
    # Create gauge chart
    fig = go.Figure(go.Indicator(
        mode="gauge+number+delta",
        value=sentiment_pct,
        domain={'x': [0, 1], 'y': [0, 1]},
        title={'text': "National Sentiment Index", 'font': {'size': 24}},
        delta={'reference': 50, 'suffix': '%'},
        gauge={
            'axis': {'range': [0, 100], 'tickwidth': 1, 'tickcolor': "white"},
            'bar': {'color': "#ff9933"},
            'bgcolor': "white",
            'borderwidth': 2,
            'bordercolor': "gray",
            'steps': [
                {'range': [0, 33], 'color': '#ef4444'},
                {'range': [33, 66], 'color': '#f59e0b'},
                {'range': [66, 100], 'color': '#10b981'}
            ],
            'threshold': {
                'line': {'color': "red", 'width': 4},
                'thickness': 0.75,
                'value': sentiment_pct
            }
        }
    ))
    
    fig.update_layout(
        height=400,
        template='plotly_dark',
        font={'size': 16}
    )
    
    st.plotly_chart(fig, use_container_width=True, key='sentiment_gauge_india')
    
    # Sentiment breakdown
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.markdown(
            f"""<div style="background:#ef4444; padding:20px; border-radius:10px; 
            text-align:center; color:white;">
            <div style="font-size:32px;">😟</div>
            <div style="font-size:14px; margin-top:10px;">CONCERNED</div>
            <div style="font-size:24px; font-weight:700;">{100-sentiment_pct if sentiment_pct < 50 else 0}%</div>
            </div>""",
            unsafe_allow_html=True
        )
    
    with col2:
        st.markdown(
            f"""<div style="background:#f59e0b; padding:20px; border-radius:10px; 
            text-align:center; color:white;">
            <div style="font-size:32px;">😐</div>
            <div style="font-size:14px; margin-top:10px;">NEUTRAL</div>
            <div style="font-size:24px; font-weight:700;">{50 if 40 <= sentiment_pct <= 60 else 0}%</div>
            </div>""",
            unsafe_allow_html=True
        )
    
    with col3:
        st.markdown(
            f"""<div style="background:#10b981; padding:20px; border-radius:10px; 
            text-align:center; color:white;">
            <div style="font-size:32px;">😊</div>
            <div style="font-size:14px; margin-top:10px;">POSITIVE</div>
            <div style="font-size:24px; font-weight:700;">{sentiment_pct if sentiment_pct > 50 else 0}%</div>
            </div>""",
            unsafe_allow_html=True
        )
    
    # Vocal tone and visual background
    st.markdown("<br>", unsafe_allow_html=True)
    
    col_tone, col_visual = st.columns(2)
    
    with col_tone:
        vocal_tone = intelligence.get('vocal_tone', 'N/A')
        st.markdown(
            f"""<div style="background:#1e293b; padding:20px; border-radius:10px; 
            border-left:4px solid #8b5cf6;">
            <strong style="color:#a78bfa; font-size:16px;">🎤 Vocal Tone:</strong><br>
            <span style="color:#e2e8f0; font-size:18px;">{vocal_tone}</span>
            </div>""",
            unsafe_allow_html=True
        )
    
    with col_visual:
        visual_bg = intelligence.get('visual_background_prompt', 'N/A')
        st.markdown(
            f"""<div style="background:#1e293b; padding:20px; border-radius:10px; 
            border-left:4px solid #3b82f6;">
            <strong style="color:#60a5fa; font-size:16px;">🎨 Visual Background:</strong><br>
            <span style="color:#e2e8f0; font-size:14px;">{visual_bg}</span>
            </div>""",
            unsafe_allow_html=True
        )


def render_sentiment_distribution_wheel_india(google_data, twitter_data):
    """Sentiment distribution pie chart for India"""
    st.markdown("### 🎭 Sentiment Distribution Wheel")
    st.caption("Today's emotional landscape in India")
    
    google_df = pd.DataFrame(google_data)
    twitter_df = pd.DataFrame(twitter_data)
    
    sentiments = []
    
    if 'public_sentiment' in google_df.columns:
        sentiments.extend(google_df['public_sentiment'].tolist())
    
    if 'primary_sentiment' in twitter_df.columns:
        sentiments.extend(twitter_df['primary_sentiment'].tolist())
    elif 'sentiment' in twitter_df.columns:
        sentiments.extend(twitter_df['sentiment'].tolist())
    
    sentiment_counts = Counter([s for s in sentiments if s and s != ''])
    
    if not sentiment_counts:
        st.info("No sentiment data available")
        return
    
    labels = list(sentiment_counts.keys())
    values = list(sentiment_counts.values())
    
    color_map = {
        'excited': '#f59e0b',
        'celebrating': '#10b981',
        'curious': '#6366f1',
        'concerned': '#ef4444',
        'controversial': '#ec4899'
    }
    colors = [color_map.get(label, '#9ca3af') for label in labels]
    
    fig = go.Figure(data=[go.Pie(
        labels=[l.title() for l in labels],
        values=values,
        hole=0.4,
        marker=dict(colors=colors),
        textinfo='label+percent',
        hovertemplate='<b>%{label}</b><br>Count: %{value}<br>%{percent}<extra></extra>'
    )])
    
    fig.update_layout(
        title="🇮🇳 India Sentiment Distribution",
        showlegend=True,
        height=450,
        margin=dict(l=20, r=20, t=60, b=20)
    )
    
    st.plotly_chart(fig, use_container_width=True, key='sentiment_wheel_india')
    
    dominant = sentiment_counts.most_common(1)[0]
    st.caption(f"💡 **Dominant Mood:** {dominant[0].title()} ({int((dominant[1]/sum(values))*100)}%)")


def render_excel_export_section(supabase, analysis_date):
    """
    Export top 5 Google + top 5 Twitter trends as Excel for Canva
    """
    st.markdown("### 📊 Canva Export - Top 5 Trends")
    st.caption("Download Excel file formatted for direct Canva import")
    
    if st.button("📥 Generate Excel Export", type="primary", use_container_width=True):
        with st.spinner("🔄 Preparing export..."):
            try:
                # Fetch top 5 Google + Twitter trends
                google_result = supabase.table('google_trends')\
                    .select('*')\
                    .eq('collection_date', analysis_date)\
                    .eq('region', 'India')\
                    .order('search_volume', desc=True)\
                    .limit(5)\
                    .execute()
                
                twitter_result = supabase.table('twitter_trends')\
                    .select('*')\
                    .eq('collection_date', analysis_date)\
                    .eq('region', 'India')\
                    .order('mention_volume', desc=True)\
                    .limit(5)\
                    .execute()
                
                google_data = google_result.data if google_result.data else []
                twitter_data = twitter_result.data if twitter_result.data else []
                
                if google_data or twitter_data:
                    excel_file = create_canva_excel(google_data, twitter_data, analysis_date)
                    
                    if excel_file:
                        st.success("✅ Excel file generated and saved!")
                        st.info(f"📁 Saved to: {excel_file}")
                        
                        # Download button
                        with open(excel_file, 'rb') as f:
                            st.download_button(
                                label="📥 Download Excel for Canva",
                                data=f,
                                file_name=f"India_Trends_{analysis_date}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True
                            )
                        
                        # Preview
                        st.markdown("#### 📋 Preview:")
                        preview_df = pd.read_excel(excel_file)
                        st.dataframe(preview_df, use_container_width=True)
                    else:
                        st.error("❌ Failed to create Excel file")
                else:
                    st.warning("⚠️ No trend data found for this date")
                    
            except Exception as e:
                st.error(f"❌ Export failed: {str(e)}")


def create_canva_excel(google_data, twitter_data, analysis_date):
    """
    Create Excel with 2 sheets:
    Sheet 1: Top 5 Google Trends
    Sheet 2: Top 5 Twitter Trends
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        import os
        
        # Create workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # === SHEET 1: GOOGLE TRENDS ===
        ws_google = wb.create_sheet("Google Search Trends")
        
        # Header row
        google_headers = ['Keyword', 'Metric', 'Context', 'Why Trending', 'Sentiment']
        for col, header in enumerate(google_headers, 1):
            cell = ws_google.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, size=12, color="FFFFFF")
            cell.fill = PatternFill(start_color="4285F4", end_color="4285F4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                bottom=Side(style='thick', color='000000')
            )
        
        # Google data rows
        for idx, trend in enumerate(google_data[:5], start=2):
            # Keyword
            ws_google.cell(row=idx, column=1, value=trend.get('keyword', 'N/A'))
            
            # Metric (formatted volume)
            volume = trend.get('search_volume', 0)
            if volume >= 1_000_000:
                metric = f"{volume/1_000_000:.1f}M searches"
            elif volume >= 1000:
                metric = f"{volume/1000:.0f}K searches"
            else:
                metric = f"{volume} searches"
            ws_google.cell(row=idx, column=2, value=metric)
            
            # Context
            context = trend.get('context', 'N/A')
            ws_google.cell(row=idx, column=3, value=context)
            
            # Why Trending
            why = trend.get('why_trending', 'N/A')
            ws_google.cell(row=idx, column=4, value=why)
            
            # Sentiment
            sentiment = trend.get('public_sentiment', 'curious')
            ws_google.cell(row=idx, column=5, value=sentiment.title())
            
            # Alternate row colors
            if idx % 2 == 0:
                fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
                for col in range(1, 6):
                    ws_google.cell(row=idx, column=col).fill = fill
        
        # Column widths for Google sheet
        ws_google.column_dimensions['A'].width = 30
        ws_google.column_dimensions['B'].width = 18
        ws_google.column_dimensions['C'].width = 50
        ws_google.column_dimensions['D'].width = 50
        ws_google.column_dimensions['E'].width = 15
        
        # Wrap text for Context and Why columns
        for row in range(2, len(google_data[:5]) + 2):
            ws_google.cell(row=row, column=3).alignment = Alignment(wrap_text=True, vertical='top')
            ws_google.cell(row=row, column=4).alignment = Alignment(wrap_text=True, vertical='top')
        
        # === SHEET 2: TWITTER TRENDS ===
        ws_twitter = wb.create_sheet("Twitter Trends")
        
        # Header row
        twitter_headers = ['Hashtag', 'Mentions', 'Context', 'Why Trending', 'Sentiment']
        for col, header in enumerate(twitter_headers, 1):
            cell = ws_twitter.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, size=12, color="FFFFFF")
            cell.fill = PatternFill(start_color="1DA1F2", end_color="1DA1F2", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                bottom=Side(style='thick', color='000000')
            )
        
        # Twitter data rows
        for idx, trend in enumerate(twitter_data[:5], start=2):
            # Hashtag (ensure it starts with #)
            keyword = trend.get('keyword', 'N/A')
            if not keyword.startswith('#') and keyword != 'N/A':
                keyword = f"#{keyword}"
            ws_twitter.cell(row=idx, column=1, value=keyword)
            
            # Mentions (formatted volume)
            mentions = trend.get('mention_volume', 0)
            if mentions >= 1_000_000:
                mention_str = f"{mentions/1_000_000:.1f}M mentions"
            elif mentions >= 1000:
                mention_str = f"{mentions/1000:.0f}K mentions"
            else:
                mention_str = f"{mentions} mentions"
            ws_twitter.cell(row=idx, column=2, value=mention_str)
            
            # Context
            context = trend.get('context', 'N/A')
            ws_twitter.cell(row=idx, column=3, value=context)
            
            # Why Trending
            why = trend.get('why_trending', 'N/A')
            ws_twitter.cell(row=idx, column=4, value=why)
            
            # Sentiment
            sentiment = trend.get('primary_sentiment') or trend.get('sentiment', 'curious')
            ws_twitter.cell(row=idx, column=5, value=sentiment.title())
            
            # Alternate row colors
            if idx % 2 == 0:
                fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
                for col in range(1, 6):
                    ws_twitter.cell(row=idx, column=col).fill = fill
        
        # Column widths for Twitter sheet
        ws_twitter.column_dimensions['A'].width = 30
        ws_twitter.column_dimensions['B'].width = 18
        ws_twitter.column_dimensions['C'].width = 50
        ws_twitter.column_dimensions['D'].width = 50
        ws_twitter.column_dimensions['E'].width = 15
        
        # Wrap text for Context and Why columns
        for row in range(2, len(twitter_data[:5]) + 2):
            ws_twitter.cell(row=row, column=3).alignment = Alignment(wrap_text=True, vertical='top')
            ws_twitter.cell(row=row, column=4).alignment = Alignment(wrap_text=True, vertical='top')
        
        # Save to output directory
        output_dir = r"D:\THE FEEDROOM\Daily Summary"
        os.makedirs(output_dir, exist_ok=True)
        
        output_path = os.path.join(output_dir, f"India_Trends_{analysis_date}.xlsx")
        wb.save(output_path)
        
        return output_path
        
    except Exception as e:
        st.error(f"Excel creation error: {str(e)}")
        return None
    
def render_raw_data_tables_india(google_data, twitter_data):
    """Display raw data tables for India only with cleaned columns"""
    st.markdown("### 📋 Raw Data Tables")
    st.caption("Complete trend data for India")
    
    tab_g, tab_t = st.tabs(["🔍 Google Trends", "🐦 Twitter Trends"])
    
    with tab_g:
        if google_data and len(google_data) > 0:
            google_df = pd.DataFrame(google_data)
            
            # Select only relevant columns for Google
            columns_to_show = [
                'rank',
                'keyword',
                'search_volume',
                'category',
                'velocity',
                'context',
                'why_trending',
                'public_sentiment',
                'sentiment_score'
            ]
            
            # Filter to only existing columns
            available_columns = [col for col in columns_to_show if col in google_df.columns]
            
            if available_columns:
                display_df = google_df[available_columns].copy()
                
                # Rename columns for better readability
                column_rename = {
                    'rank': 'Rank',
                    'keyword': 'Keyword',
                    'search_volume': 'Search Volume',
                    'category': 'Category',
                    'velocity': 'Velocity',
                    'context': 'Context',
                    'why_trending': 'Why Trending',
                    'public_sentiment': 'Sentiment',
                    'sentiment_score': 'Sentiment Score'
                }
                
                display_df = display_df.rename(columns=column_rename)
                
                st.dataframe(display_df, use_container_width=True, height=400)
                st.caption(f"📊 Total: {len(display_df)} Google trends")
            else:
                st.info("No Google data columns available")
        else:
            st.info("No Google data available")
    
    with tab_t:
        if twitter_data and len(twitter_data) > 0:
            twitter_df = pd.DataFrame(twitter_data)
            
            # Select only relevant columns for Twitter
            columns_to_show = [
                'rank',
                'keyword',
                'mention_volume',
                'category',
                'velocity',
                'sentiment',
                'context',
                'why_trending'
            ]
            
            # Filter to only existing columns
            available_columns = [col for col in columns_to_show if col in twitter_df.columns]
            
            if available_columns:
                display_df = twitter_df[available_columns].copy()
                
                # Rename columns for better readability
                column_rename = {
                    'rank': 'Rank',
                    'keyword': 'Keyword',
                    'mention_volume': 'Mention Volume',
                    'category': 'Category',
                    'velocity': 'Velocity',
                    'sentiment': 'Sentiment',
                    'context': 'Context',
                    'why_trending': 'Why Trending'
                }
                
                display_df = display_df.rename(columns=column_rename)
                
                st.dataframe(display_df, use_container_width=True, height=400)
                st.caption(f"📊 Total: {len(display_df)} Twitter trends")
            else:
                st.info("No Twitter data columns available")
        else:
            st.info("No Twitter data available")